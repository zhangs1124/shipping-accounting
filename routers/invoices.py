from fastapi import APIRouter, Depends, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from typing import Optional
from datetime import date
import csv
import io
import re
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import models
from database import get_db

router = APIRouter(prefix="/invoices", tags=["invoices"])
templates = Jinja2Templates(directory="templates")

INVOICE_STATUSES = ["草稿", "已開立", "已收款"]


def generate_invoice_no(db: Session, invoice_date: date) -> str:
    prefix = f"A{invoice_date.strftime('%Y%m%d')}-"
    pattern = re.compile(rf"^A{invoice_date.strftime('%Y%m%d')}\-(\d{{3}})$")
    existing_nos = (
        db.query(models.Invoice.invoice_no)
        .filter(models.Invoice.invoice_no.like(f"{prefix}%"))
        .all()
    )

    max_seq = 0
    for (invoice_no,) in existing_nos:
        m = pattern.match(invoice_no)
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return f"{prefix}{max_seq + 1:03d}"


@router.get("", response_class=HTMLResponse)
def list_invoices(
    request: Request,
    voyage_id: Optional[int] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    invoices = _query_invoices(db, voyage_id, status, date_from, date_to)
    voyages = db.query(models.Voyage).order_by(models.Voyage.voyage_no).all()

    return templates.TemplateResponse("invoices/list.html", {
        "request": request,
        "invoices": invoices,
        "voyages": voyages,
        "statuses": INVOICE_STATUSES,
        "filter": {
            "voyage_id": voyage_id,
            "status": status,
            "date_from": date_from,
            "date_to": date_to,
        },
    })

def _query_invoices(
    db: Session,
    voyage_id: Optional[int],
    status: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
):
    query = db.query(models.Invoice)
    if voyage_id:
        query = query.filter(models.Invoice.voyage_id == voyage_id)
    if status:
        query = query.filter(models.Invoice.status == status)
    if date_from:
        query = query.filter(models.Invoice.invoice_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(models.Invoice.invoice_date <= date.fromisoformat(date_to))
    return query.order_by(models.Invoice.invoice_date.desc()).all()


@router.get("/report/print", response_class=HTMLResponse)
def print_invoice_report(
    request: Request,
    voyage_id: Optional[int] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    invoices = _query_invoices(db, voyage_id, status, date_from, date_to)
    voyages = db.query(models.Voyage).order_by(models.Voyage.voyage_no).all()
    total_amount = sum((inv.total_amount for inv in invoices), Decimal("0"))
    return templates.TemplateResponse("invoices/list.html", {
        "request": request,
        "invoices": invoices,
        "voyages": voyages,
        "statuses": INVOICE_STATUSES,
        "print_mode": True,
        "print_generated_at": date.today().strftime("%Y-%m-%d"),
        "report_total_amount": total_amount,
        "filter": {
            "voyage_id": voyage_id,
            "status": status,
            "date_from": date_from,
            "date_to": date_to,
        },
    })


@router.get("/report/export-excel")
def export_invoice_report_excel(
    voyage_id: Optional[int] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    invoices = _query_invoices(db, voyage_id, status, date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "帳務報表"

    ws["A1"] = "船務部帳務系統 - 帳務報表"
    ws["A2"] = f"列印日期：{date.today().strftime('%Y-%m-%d')}"
    ws["A3"] = f"篩選條件：航次={voyage_id or '全部'}、狀態={status or '全部'}、日期={date_from or '-'}~{date_to or '-'}"
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    ws.merge_cells("A3:J3")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    label_fill = PatternFill("solid", fgColor="F0F4F8")
    header_fill = PatternFill("solid", fgColor="1A6FC4")
    white_bold_font = Font(bold=True, color="FFFFFF")

    row = 5
    total_amount = Decimal("0")
    for inv in invoices:
        # 主檔區塊
        ws.cell(row=row, column=1, value="帳單主檔").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = label_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1

        master_pairs = [
            ("帳單編號", inv.invoice_no, "航次", inv.voyage.voyage_no),
            ("客戶名稱", inv.customer_name, "帳單日期", inv.invoice_date.strftime("%Y-%m-%d")),
            ("帳單狀態", inv.status, "建立時間", inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else ""),
        ]
        for left_label, left_value, right_label, right_value in master_pairs:
            ws.cell(row=row, column=1, value=left_label).fill = label_fill
            ws.cell(row=row, column=2, value=left_value)
            ws.cell(row=row, column=4, value=right_label).fill = label_fill
            ws.cell(row=row, column=5, value=right_value)
            for c in (1, 2, 4, 5):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        row += 1
        # 明細區塊
        detail_headers = ["#", "項目代碼", "項目名稱", "數量", "單價", "幣別", "小計", "備註"]
        for i, h in enumerate(detail_headers, start=1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = white_bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        row += 1

        if inv.lines:
            for idx, line in enumerate(inv.lines, start=1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=line.charge_item.code)
                ws.cell(row=row, column=3, value=line.charge_item.name)
                qty_cell = ws.cell(row=row, column=4, value=float(line.quantity))
                unit_cell = ws.cell(row=row, column=5, value=float(line.unit_price))
                ws.cell(row=row, column=6, value=line.currency)
                subtotal_cell = ws.cell(row=row, column=7, value=float(line.subtotal))
                ws.cell(row=row, column=8, value=line.remark or "")
                qty_cell.number_format = '#,##0.####'
                unit_cell.number_format = '#,##0.00'
                subtotal_cell.number_format = '#,##0.00'
                for c in range(1, 9):
                    ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=8).alignment = Alignment(horizontal="left", vertical="center")
                for c in (1, 2, 4, 5, 6, 7):
                    ws.cell(row=row, column=c).alignment = Alignment(horizontal="center" if c in (1, 2, 6) else "right", vertical="center")
                row += 1
        else:
            ws.cell(row=row, column=1, value="無明細")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=1).border = thin_border
            row += 1

        ws.cell(row=row, column=6, value="帳單總計").font = Font(bold=True)
        ws.cell(row=row, column=6).fill = label_fill
        invoice_total_cell = ws.cell(row=row, column=7, value=float(inv.total_amount))
        invoice_total_cell.font = Font(bold=True)
        invoice_total_cell.number_format = '#,##0.00'
        invoice_total_cell.alignment = Alignment(horizontal="right")
        for c in (6, 7):
            ws.cell(row=row, column=c).border = thin_border
        row += 2

        total_amount += inv.total_amount

    ws.cell(row=row, column=6, value="報表整體總計").font = Font(bold=True)
    ws.cell(row=row, column=6).fill = label_fill
    total_cell = ws.cell(row=row, column=7, value=float(total_amount))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    for c in (6, 7):
        ws.cell(row=row, column=c).border = thin_border

    widths = {
        "A": 12, "B": 14, "C": 28, "D": 12, "E": 12, "F": 12, "G": 14,
        "H": 24, "I": 12, "J": 12
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for r in range(5, row + 1):
        ws.row_dimensions[r].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"invoice_report_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/new", response_class=HTMLResponse)
def new_invoice_form(request: Request, db: Session = Depends(get_db)):
    voyages = db.query(models.Voyage).order_by(models.Voyage.voyage_no).all()
    today = date.today()
    return templates.TemplateResponse("invoices/new.html", {
        "request": request,
        "voyages": voyages,
        "suggested_invoice_no": generate_invoice_no(db, today),
        "today_str": today.isoformat(),
    })


@router.post("")
def create_invoice(
    request: Request,
    voyage_id: int = Form(...),
    customer_name: str = Form(...),
    invoice_date: str = Form(...),
    db: Session = Depends(get_db),
):
    invoice_dt = date.fromisoformat(invoice_date)
    for _ in range(5):
        invoice_no = generate_invoice_no(db, invoice_dt)
        invoice = models.Invoice(
            invoice_no=invoice_no,
            voyage_id=voyage_id,
            customer_name=customer_name,
            invoice_date=invoice_dt,
            status="草稿",
            total_amount=0,
        )
        db.add(invoice)
        try:
            db.commit()
            db.refresh(invoice)
            return RedirectResponse(url=f"/invoices/{invoice.id}", status_code=303)
        except IntegrityError:
            # Unique key collision can happen when multiple users create at the same time.
            db.rollback()
            continue

    voyages = db.query(models.Voyage).order_by(models.Voyage.voyage_no).all()
    return templates.TemplateResponse("invoices/new.html", {
        "request": request,
        "voyages": voyages,
        "error": "系統忙碌中，請稍後重試建立帳單",
        "form_data": {
            "voyage_id": voyage_id,
            "customer_name": customer_name,
            "invoice_date": invoice_date,
        },
        "suggested_invoice_no": generate_invoice_no(db, invoice_dt),
        "today_str": date.today().isoformat(),
    }, status_code=409)


@router.get("/{invoice_id}", response_class=HTMLResponse)
def invoice_detail(invoice_id: int, request: Request, db: Session = Depends(get_db)):
    invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not invoice:
        return RedirectResponse(url="/invoices", status_code=303)

    charge_items = db.query(models.ChargeItem).order_by(models.ChargeItem.code).all()
    return templates.TemplateResponse("invoices/detail.html", {
        "request": request,
        "invoice": invoice,
        "charge_items": charge_items,
        "statuses": INVOICE_STATUSES,
        "can_edit": invoice.status == "草稿",
    })


@router.post("/{invoice_id}/status")
def update_invoice_status(
    invoice_id: int,
    status: str = Form(...),
    db: Session = Depends(get_db),
):
    invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not invoice:
        return RedirectResponse(url="/invoices", status_code=303)

    if status in INVOICE_STATUSES:
        invoice.status = status
        db.commit()
    return RedirectResponse(url=f"/invoices/{invoice_id}", status_code=303)


@router.post("/{invoice_id}/delete")
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not invoice:
        return RedirectResponse(url="/invoices?error=帳單不存在", status_code=303)

    if invoice.status != "草稿":
        return RedirectResponse(url=f"/invoices/{invoice_id}?error=僅草稿狀態可刪除主檔", status_code=303)

    if invoice.lines:
        return RedirectResponse(url=f"/invoices/{invoice_id}?error=請先刪除所有明細，才可刪除主檔", status_code=303)

    db.delete(invoice)
    db.commit()
    return RedirectResponse(url="/invoices", status_code=303)


@router.get("/{invoice_id}/print", response_class=HTMLResponse)
def print_invoice(invoice_id: int, request: Request, db: Session = Depends(get_db)):
    invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not invoice:
        return RedirectResponse(url="/invoices", status_code=303)
    return templates.TemplateResponse("invoices/print.html", {
        "request": request,
        "invoice": invoice,
    })


@router.get("/{invoice_id}/export-csv")
def export_csv(invoice_id: int, db: Session = Depends(get_db)):
    invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not invoice:
        return RedirectResponse(url="/invoices", status_code=303)

    output = io.StringIO()
    writer = csv.writer(output)

    # 標頭資訊
    writer.writerow(["帳單編號", invoice.invoice_no])
    writer.writerow(["客戶名稱", invoice.customer_name])
    writer.writerow(["帳單日期", invoice.invoice_date.strftime("%Y-%m-%d")])
    writer.writerow(["航次編號", invoice.voyage.voyage_no])
    writer.writerow(["船舶", f"{invoice.voyage.ship.code} - {invoice.voyage.ship.name}"])
    writer.writerow(["狀態", invoice.status])
    writer.writerow([])

    # 明細標題
    writer.writerow(["項目代碼", "項目名稱", "數量", "單價", "幣別", "小計", "備註"])
    for line in invoice.lines:
        writer.writerow([
            line.charge_item.code,
            line.charge_item.name,
            float(line.quantity),
            float(line.unit_price),
            line.currency,
            float(line.subtotal),
            line.remark or "",
        ])
    writer.writerow([])
    writer.writerow(["", "", "", "", "總計", float(invoice.total_amount), ""])

    output.seek(0)
    filename = f"invoice_{invoice.invoice_no}.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
