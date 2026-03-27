from fastapi import APIRouter, Depends, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from utils.templates import templates
from sqlalchemy.orm import Session
from datetime import datetime
from typing import Optional
import io
import pandas as pd
import openpyxl
import openpyxl.styles as styles

import models
from database import get_db

router = APIRouter(prefix="/voyage-tasks", tags=["voyage_tasks"])


# ────────────────────────────────────────────────────────────
#  列表頁：顯示所有已建立進出港記錄的航次 + 上方表單新增
# ────────────────────────────────────────────────────────────
@router.get("", response_class=HTMLResponse)
def list_voyage_tasks(request: Request, db: Session = Depends(get_db)):
    """列出所有已建立進出港記錄的航次，並提供上方新增表單。"""
    # 取 ships 清單供 autocomplete
    ships = db.query(models.Ship).order_by(models.Ship.name).all()

    # 取所有有任務記錄的航次（distinct voyage_id）
    voyage_ids_with_logs = (
        db.query(models.VoyageTaskLog.voyage_id).distinct().all()
    )
    voyage_ids_with_logs = [r[0] for r in voyage_ids_with_logs]

    voyages_with_tasks = []
    for vid in voyage_ids_with_logs:
        voyage = db.query(models.Voyage).filter(models.Voyage.id == vid).first()
        if voyage:
            total = db.query(models.VoyageTaskLog).filter(
                models.VoyageTaskLog.voyage_id == vid
            ).count()
            done = db.query(models.VoyageTaskLog).filter(
                models.VoyageTaskLog.voyage_id == vid,
                models.VoyageTaskLog.recorded_time.isnot(None)
            ).count()
            voyages_with_tasks.append({
                "voyage": voyage,
                "total": total,
                "done": done,
            })

    # 依照航次建立時間倒序排列
    voyages_with_tasks.sort(key=lambda x: x["voyage"].created_at or datetime.min, reverse=True)

    return templates.TemplateResponse("voyage_tasks/list.html", {
        "request": request,
        "ships": ships,
        "voyages_with_tasks": voyages_with_tasks,
    })


# ────────────────────────────────────────────────────────────
#  API：取得某艘船的航次清單
# ────────────────────────────────────────────────────────────
@router.get("/api/voyages")
def api_get_voyages(ship_name: str, db: Session = Depends(get_db)):
    ship = db.query(models.Ship).filter(
        (models.Ship.name == ship_name) | (models.Ship.code == ship_name)
    ).first()
    if not ship:
        return []
    
    voyages = db.query(models.Voyage).filter(models.Voyage.ship_id == ship.id).order_by(models.Voyage.voyage_no.desc()).all()
    return [{"id": v.id, "voyage_no": v.voyage_no} for v in voyages]


# ────────────────────────────────────────────────────────────
#  API：取得所有船名（autocomplete 用）
# ────────────────────────────────────────────────────────────
@router.get("/api/ships")
def api_get_ships(db: Session = Depends(get_db)):
    ships = db.query(models.Ship).order_by(models.Ship.name).all()
    return [{"id": s.id, "name": s.name, "code": s.code} for s in ships]


# ────────────────────────────────────────────────────────────
#  建立新的進出港記錄：依船名找船 → 依航次編號找 voyage → 初始化
# ────────────────────────────────────────────────────────────
@router.post("/create")
def create_voyage_task(
    ship_name: str = Form(...),
    voyage_no: str = Form(...),
    db: Session = Depends(get_db)
):
    """接收船名與航次編號，找到對應的 Voyage，並初始化所有啟用的 TaskCategory。"""
    # 找船（模糊匹配名稱或代碼）
    ship = db.query(models.Ship).filter(
        (models.Ship.name == ship_name) | (models.Ship.code == ship_name)
    ).first()
    if not ship:
        return RedirectResponse(
            url=f"/voyage-tasks?error=找不到船舶：{ship_name}",
            status_code=303
        )

    # 找航次
    voyage = db.query(models.Voyage).filter(
        models.Voyage.voyage_no == voyage_no,
        models.Voyage.ship_id == ship.id
    ).first()
    if not voyage:
        return RedirectResponse(
            url=f"/voyage-tasks?error=找不到航次：{voyage_no}（船：{ship_name}）",
            status_code=303
        )

    # 檢查是否已經初始化過
    existing_count = db.query(models.VoyageTaskLog).filter(
        models.VoyageTaskLog.voyage_id == voyage.id
    ).count()

    if existing_count == 0:
        # 初始化所有啟用中的 TaskCategory
        categories = db.query(models.TaskCategory).filter(
            models.TaskCategory.is_active == 1
        ).order_by(
            models.TaskCategory.display_order,
            models.TaskCategory.task_group,
            models.TaskCategory.name
        ).all()
        for cat in categories:
            new_log = models.VoyageTaskLog(
                voyage_id=voyage.id,
                task_id=cat.id,
                recorded_time=None,
                recorded_by="",
                remarks=""
            )
            db.add(new_log)
        db.commit()

    return RedirectResponse(url=f"/voyage-tasks/{voyage.id}", status_code=303)


# ────────────────────────────────────────────────────────────
#  詳細頁：顯示某航次的所有檢查項目 + 支援加減
# ────────────────────────────────────────────────────────────
@router.get("/{voyage_id}", response_class=HTMLResponse)
def detail_voyage_tasks(
    voyage_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    voyage = db.query(models.Voyage).filter(models.Voyage.id == voyage_id).first()
    if not voyage:
        return RedirectResponse(url="/voyage-tasks", status_code=303)

    # 讀取此航次的任務清單（依分組、順序排序）
    task_logs = (
        db.query(models.VoyageTaskLog)
        .join(models.TaskCategory)
        .filter(models.VoyageTaskLog.voyage_id == voyage_id)
        .order_by(
            models.TaskCategory.display_order,
            models.TaskCategory.task_group,
            models.TaskCategory.name
        )
        .all()
    )

    # 可加入的項目（尚未在清單中）
    used_task_ids = [log.task_id for log in task_logs]
    if used_task_ids:
        available_categories = (
            db.query(models.TaskCategory)
            .filter(
                models.TaskCategory.is_active == 1,
                models.TaskCategory.id.notin_(used_task_ids)
            )
            .order_by(models.TaskCategory.task_group, models.TaskCategory.display_order)
            .all()
        )
    else:
        available_categories = (
            db.query(models.TaskCategory)
            .filter(models.TaskCategory.is_active == 1)
            .order_by(models.TaskCategory.task_group, models.TaskCategory.display_order)
            .all()
        )

    return templates.TemplateResponse("voyage_tasks/detail.html", {
        "request": request,
        "voyage": voyage,
        "task_logs": task_logs,
        "available_categories": available_categories,
    })


# ────────────────────────────────────────────────────────────
#  列印頁：顯示格式化報表供列印
# ────────────────────────────────────────────────────────────
@router.get("/{voyage_id}/print", response_class=HTMLResponse)
def print_voyage_tasks(
    voyage_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    voyage = db.query(models.Voyage).filter(models.Voyage.id == voyage_id).first()
    if not voyage:
        return RedirectResponse(url="/voyage-tasks", status_code=303)

    # 讀取清單（依自定義顯示順序排序）
    task_logs = (
        db.query(models.VoyageTaskLog)
        .join(models.TaskCategory)
        .filter(models.VoyageTaskLog.voyage_id == voyage_id)
        .order_by(
            models.TaskCategory.display_order,
            models.TaskCategory.task_group,
            models.TaskCategory.name
        )
        .all()
    )

    return templates.TemplateResponse("voyage_tasks/print.html", {
        "request": request,
        "voyage": voyage,
        "task_logs": task_logs,
        "print_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })


# ────────────────────────────────────────────────────────────
#  Excel 匯出：產生格式化的 Excel 檔案供下載
# ────────────────────────────────────────────────────────────
@router.get("/{voyage_id}/excel")
def export_voyage_tasks_excel(
    voyage_id: int,
    db: Session = Depends(get_db)
):
    voyage = db.query(models.Voyage).filter(models.Voyage.id == voyage_id).first()
    if not voyage:
        return RedirectResponse(url="/voyage-tasks", status_code=303)

    # 讀取清單（依自定義顯示順序排序）
    task_logs = (
        db.query(models.VoyageTaskLog)
        .join(models.TaskCategory)
        .filter(models.VoyageTaskLog.voyage_id == voyage_id)
        .order_by(
            models.TaskCategory.display_order,
            models.TaskCategory.task_group,
            models.TaskCategory.name
        )
        .all()
    )

    # 準備資料
    data = []
    for i, log in enumerate(task_logs, 1):
        data.append({
            "#": i,
            "任務分組": log.task_category.task_group or '未分類',
            "任務項目內容": log.task_category.name,
            "執行時間": log.recorded_time.strftime('%Y-%m-%d %H:%M') if log.recorded_time else '尚未執行',
            "備註": log.remarks or ''
        })

    # 使用 io.BytesIO 作為記憶體緩衝區
    output = io.BytesIO()
    
    # 建立一個新的 Excel 活頁簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "進出港任務"
    
    # 1. 寫入主檔資訊 (標題與基礎資訊)
    ws.merge_cells('A1:E1')
    ws['A1'] = f"進出港任務報表 - {voyage.voyage_no}"
    ws['A1'].font = styles.Font(bold=True, size=16, color="1A3A5C")
    ws['A1'].alignment = styles.Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f"船舶名稱: {voyage.ship.name} ({voyage.ship.code})"
    ws['C2'] = f"航次編號: {voyage.voyage_no}"
    ws['A3'] = f"裝卸港口: {voyage.port_of_loading or '-'} ➔ {voyage.port_of_discharge or '-'}"
    ws['C3'] = f"匯出日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # 子標題字體
    sub_font = styles.Font(bold=True)
    ws['A2'].font = sub_font
    ws['C2'].font = sub_font
    ws['A3'].font = sub_font
    ws['C3'].font = sub_font
    
    # 2. 寫入表格標題
    headers = ["#", "任務分組", "任務項目內容", "執行時間", "備註"]
    header_font = styles.Font(bold=True, color="FFFFFF")
    header_fill = styles.PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = styles.Alignment(horizontal='center')
        
    # 3. 寫入資料內容
    thin_border = styles.Side(border_style="thin", color="000000")
    border = styles.Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
    
    for r_idx, row_data in enumerate(data, 6):
        ws.cell(row=r_idx, column=1, value=row_data["#"]).alignment = styles.Alignment(horizontal='center')
        ws.cell(row=r_idx, column=2, value=row_data["任務分組"]).alignment = styles.Alignment(horizontal='center')
        ws.cell(row=r_idx, column=3, value=row_data["任務項目內容"])
        ws.cell(row=r_idx, column=4, value=row_data["執行時間"]).alignment = styles.Alignment(horizontal='center')
        ws.cell(row=r_idx, column=5, value=row_data["備註"])
        
        # 套用框線
        for c_idx in range(1, 6):
            ws.cell(row=r_idx, column=c_idx).border = border
            
    # 對標題列也套用框線
    for c_idx in range(1, 6):
        ws.cell(row=5, column=c_idx).border = border
    
    # 4. 欄位寬度調整
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35

    wb.save(output)
    output.seek(0)
    
    # 下載檔名
    safe_voyage_no = voyage.voyage_no.replace("/", "_").replace("\\", "_")
    filename = f"VoyageTasks_{safe_voyage_no}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ────────────────────────────────────────────────────────────
#  API：新增任務項目到此航次
# ────────────────────────────────────────────────────────────
@router.post("/api/add")
def api_add_task_log(
    voyage_id: int = Form(...),
    task_id: int = Form(...),
    db: Session = Depends(get_db)
):
    """將選取的任務加入此航次的清單中"""
    existing = db.query(models.VoyageTaskLog).filter(
        models.VoyageTaskLog.voyage_id == voyage_id,
        models.VoyageTaskLog.task_id == task_id
    ).first()
    if existing:
        return JSONResponse({"error": "該項目已經存在於清單中"}, status_code=400)

    cat = db.query(models.TaskCategory).filter(models.TaskCategory.id == task_id).first()
    if not cat:
        return JSONResponse({"error": "找不到此任務類別"}, status_code=404)

    new_log = models.VoyageTaskLog(
        voyage_id=voyage_id,
        task_id=task_id,
        recorded_time=None,
        recorded_by="",
        remarks=""
    )
    db.add(new_log)
    db.commit()
    db.refresh(new_log)
    return {
        "status": "success",
        "id": new_log.id,
        "task_name": cat.name,
        "task_group": cat.task_group or "未分類",
        "default_fee": float(cat.default_fee or 0),
    }


# ────────────────────────────────────────────────────────────
#  API：更新紀錄時間與備註
# ────────────────────────────────────────────────────────────
@router.post("/api/update/{log_id}")
def api_update_task_log(
    log_id: int,
    recorded_time: Optional[str] = Form(None),
    remarks: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    log = db.query(models.VoyageTaskLog).filter(models.VoyageTaskLog.id == log_id).first()
    if not log:
        return JSONResponse({"error": "紀錄不存在"}, status_code=404)

    log_time = None
    if recorded_time and recorded_time.strip():
        try:
            log_time = datetime.fromisoformat(recorded_time.replace("T", " "))
        except ValueError:
            try:
                log_time = datetime.strptime(recorded_time, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return JSONResponse({"error": "時間格式錯誤"}, status_code=400)

    log.recorded_time = log_time
    if remarks is not None:
        log.remarks = remarks
        
    if log_time:
        # 當任務完成時，自動結案所有相關之未結案提醒
        open_reminders = db.query(models.Reminder).filter(
            models.Reminder.source_table == "voyage_task_logs",
            models.Reminder.source_id == log_id,
            models.Reminder.is_closed == 0
        ).all()
        for r in open_reminders:
            r.is_closed = 1
            r.updated_at = datetime.now()

    db.commit()
    return {"status": "success"}


# ────────────────────────────────────────────────────────────
#  API：移除此航次的某個任務項目
# ────────────────────────────────────────────────────────────
@router.post("/api/delete/{log_id}")
def api_delete_task_log(
    log_id: int,
    db: Session = Depends(get_db)
):
    log = db.query(models.VoyageTaskLog).filter(models.VoyageTaskLog.id == log_id).first()
    if not log:
        return JSONResponse({"error": "紀錄不存在"}, status_code=404)

    db.delete(log)
    db.commit()
    return {"status": "success"}


# ────────────────────────────────────────────────────────────
#  API：刪除整筆航次的所有進出港任務紀錄
# ────────────────────────────────────────────────────────────
@router.post("/{voyage_id}/purge")
def purge_voyage_tasks(voyage_id: int, db: Session = Depends(get_db)):
    """刪除特定航次的所有進出港任務紀錄。"""
    from utils.audit_logger import log_action
    
    # 紀錄刪除動作
    log_action(
        db, 
        action="DELETE_ALL_TASKS", 
        table_name="voyage_task_logs", 
        target_id=str(voyage_id),
        new_value={"message": "Purged all task logs for voyage and related reminders"}
    )
    
    db.query(models.VoyageTaskLog).filter(
        models.VoyageTaskLog.voyage_id == voyage_id
    ).delete(synchronize_session=False)
    
    db.commit()
    return RedirectResponse(url="/voyage-tasks", status_code=303)
